# Task 1 Script - Read the data from the two provided excel sheets and
# store the data in two corresponding database tables.

import openpyxl
import sqlite3

# Load the two provided workbooks
print('Loading Workbooks...')
inspections_wb = openpyxl.load_workbook('data/inspections.xlsx')
violations_wb = openpyxl.load_workbook('data/violations.xlsx')
print('Workbooks loaded')

# Connect to the database
connection = sqlite3.connect('violations.db')
cursor = connection.cursor()

# Drop the two tables if they already exist
cursor.execute('DROP TABLE IF EXISTS Violations')
cursor.execute('DROP TABLE IF EXISTS Inspections')

# Create the Inspections table
print('Creating Tables...')
sql_command = '''
    CREATE TABLE Inspections (
        inspection_id INTEGER NOT NULL,
        activity_date DATE,
        employee_id CHAR(9),
        facility_address VARCHAR(255),
        facility_city VARCHAR(255),
        facility_id CHAR(9),
        facility_name VARCHAR(255),
        facility_state CHAR(2),
        facility_zip INTEGER,
        grade CHAR(1),
        owner_id CHAR(9),
        owner_name VARCHAR(255),
        pe_description VARCHAR(255),
        program_element_pe INTEGER,
        program_name VARCHAR(255),
        program_status VARCHAR(10),
        record_id CHAR(9),
        score INTEGER,
        serial_number VARCHAR(10),
        service_code INTEGER,
        service_description VARCHAR(50),
        CONSTRAINT PK_inspection_id PRIMARY KEY (inspection_id)
    )
'''

cursor.execute(sql_command)

# Create the Violations table
sql_command = '''
    CREATE TABLE Violations (
        violation_id INTEGER NOT NULL,
        points INTEGER,
        serial_number VARCHAR(10),
        violation_code CHAR(4),
        violation_description VARCHAR(255),
        violation_status varchar(30),
        CONSTRAINT PK_violation_id PRIMARY KEY (violation_id),
        CONSTRAINT FK_serial_number 
            FOREIGN KEY (serial_number) 
            REFERENCES Inspections (serial_number)
    )
'''

cursor.execute(sql_command)

print('Tables created')

# Open the active sheet in the Inspections workbook and add each row 
# of data to the corresponding columns in the Inspections table
print('Importing Data...')
sheet = inspections_wb['inspections']
for i in range(2, sheet.max_row + 1):
    sql_command = '''
        INSERT INTO Inspections (
            activity_date,
            employee_id,
            facility_address,
            facility_city,
            facility_id,
            facility_name,
            facility_state,
            facility_zip,
            grade,
            owner_id,
            owner_name,
            pe_description,
            program_element_pe,
            program_name,
            program_status,
            record_id,
            score,
            serial_number,
            service_code,
            service_description
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    '''

    cursor.execute(sql_command, (
        sheet.cell(row=i, column=1).value,
        sheet.cell(row=i, column=2).value,
        sheet.cell(row=i, column=3).value,
        sheet.cell(row=i, column=4).value,
        sheet.cell(row=i, column=5).value,
        sheet.cell(row=i, column=6).value,
        sheet.cell(row=i, column=7).value,
        # The number after the hyphen in the zip code is not needed for the reports
        int(str(sheet.cell(row=i, column=8).value).split('-')[0]),
        sheet.cell(row=i, column=9).value,
        sheet.cell(row=i, column=10).value,
        sheet.cell(row=i, column=11).value,
        sheet.cell(row=i, column=12).value,
        sheet.cell(row=i, column=13).value,
        sheet.cell(row=i, column=14).value,
        sheet.cell(row=i, column=15).value,
        sheet.cell(row=i, column=16).value,
        sheet.cell(row=i, column=17).value,
        sheet.cell(row=i, column=18).value,
        sheet.cell(row=i, column=19).value,
        sheet.cell(row=i, column=20).value
    ))

# Open the active sheet in the Violations workbook and add each row of
# data to the corresponding columns in the Violations table
sheet = violations_wb['violations']
for i in range(2, sheet.max_row + 1):
    sql_command = '''
        INSERT INTO Violations (
            points,
            serial_number,
            violation_code,
            violation_description,
            violation_status
        )
        VALUES (?,?,?,?,?)
    '''

    cursor.execute(sql_command, (
        sheet.cell(row=i, column=1).value,
        sheet.cell(row=i, column=2).value,
        sheet.cell(row=i, column=3).value,
        sheet.cell(row=i, column=4).value,
        sheet.cell(row=i, column=5).value
    ))
print('Data Imported')

# Save changes to the database and close the connection
connection.commit()
connection.close()

