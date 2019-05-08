# Task 3 Script - Creates a new excel file which stores the code, the
# description and the total count of each type of violation

import sqlite3
import openpyxl
from openpyxl.styles import Font

print('Creating Worksheet...')

# Setup connection to the database
connection = sqlite3.connect('violations.db')
cursor = connection.cursor()

# Query database for violation code, description and total count
# for each type of violation
sql_command='''
    SELECT violation_code, violation_description, COUNT(violation_code)
    FROM Violations
    GROUP BY violation_code
'''

cursor.execute(sql_command)
result = cursor.fetchall()

# Open new workbook and sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Set sheet headers
sheet.title = 'Violation Types'
sheet['A1'] = 'Code'
sheet['A1'].font = Font(bold=True)
sheet['B1'] = 'Description'
sheet['B1'].font = Font(bold=True)
sheet['C1'] = 'Count'
sheet['C1'].font = Font(bold=True)

# Add data from query result into the new sheet
column_width = 0
for i in range(0, len(result)):
    sheet.cell(row=(i+2), column=1).value = result[i][0]
    sheet.cell(row=(i+2), column=2).value = result[i][1]
    sheet.cell(row=(i+2), column=3).value = result[i][2]
    # Check that width is set so all information can be seen
    if (len(result[i][1]) > column_width):
        column_width = len(result[i][1])

# Display total amount of violations at the bottom of the sheet
sheet.cell(row=((len(result) + 2)), column=2).value = 'Total Violations'
sheet.cell(row=((len(result) + 2)), column=3).value =('=SUM(C2:C%d)' % (len(result) + 1))

# Set column B to correct width and save the new excel file
sheet.column_dimensions['B'].width = column_width
wb.save('ViolationTypes.xlsx')

# Close database connection
connection.close()

print('Worksheet Created')